"""Exportador a Excel (openpyxl).

Genera un .xlsx con:
  - Hoja "Comprobante": cabecera + items.
  - Hoja "Validación": hallazgos detectados (para revisión humana).
Pensado para que un contador lo pegue directo en su flujo o ERP.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..application.extract_document import ExtractionResult

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_LABEL_FONT = Font(bold=True)
_ERROR_FILL = PatternFill("solid", fgColor="FECACA")
_WARN_FILL = PatternFill("solid", fgColor="FEF3C7")


def export_to_excel(result: ExtractionResult) -> bytes:
    inv = result.invoice
    wb = Workbook()

    # --- Hoja 1: Comprobante ---
    ws = wb.active
    ws.title = "Comprobante"

    header = [
        ("Tipo de documento", inv.document_type.value),
        ("Número", inv.full_number or ""),
        ("Fecha de emisión", inv.issue_date.isoformat() if inv.issue_date else ""),
        ("Emisor", inv.issuer_name or ""),
        ("RUC emisor", str(inv.issuer_ruc) if inv.issuer_ruc else ""),
        ("Cliente", inv.customer_name or ""),
        ("Doc. cliente", inv.customer_doc or ""),
        ("Moneda", inv.currency),
        ("Subtotal", float(inv.subtotal)),
        ("IGV (18%)", float(inv.igv)),
        ("Total", float(inv.total)),
        ("Confianza", result.confidence),
    ]
    for r, (label, value) in enumerate(header, start=1):
        ws.cell(r, 1, label).font = _LABEL_FONT
        ws.cell(r, 2, value)

    start = len(header) + 2
    cols = ["Descripción", "Cantidad", "P. Unitario", "Total línea"]
    for c, name in enumerate(cols, start=1):
        cell = ws.cell(start, c, name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for i, it in enumerate(inv.items, start=1):
        row = start + i
        ws.cell(row, 1, it.description)
        ws.cell(row, 2, float(it.quantity))
        ws.cell(row, 3, float(it.unit_price))
        ws.cell(row, 4, float(it.line_total))

    _autosize(ws)

    # --- Hoja 2: Validación ---
    wv = wb.create_sheet("Validación")
    for c, name in enumerate(["Campo", "Severidad", "Mensaje"], start=1):
        cell = wv.cell(1, c, name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    if not result.issues:
        wv.cell(2, 1, "Sin observaciones ✓")
    else:
        for i, issue in enumerate(result.issues, start=2):
            wv.cell(i, 1, issue.field)
            sev = wv.cell(i, 2, issue.severity.value)
            wv.cell(i, 3, issue.message)
            sev.fill = _ERROR_FILL if issue.severity.value == "error" else _WARN_FILL
    _autosize(wv)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _autosize(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 3, max_width)
    for cell in ws[1]:
        cell.alignment = Alignment(vertical="center")
