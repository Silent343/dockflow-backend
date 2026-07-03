"""Reporte consolidado de un lote de comprobantes (openpyxl).

Genera un .xlsx con tres hojas:
  - Resumen: una fila por archivo, coloreada por estado.
  - Validación: todos los hallazgos de todos los archivos.
  - Totales: conteos y montos agregados.
"""
from __future__ import annotations

import io
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from ..batch import BatchRow

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_STATUS_FILL = {
    "OK": PatternFill("solid", fgColor="BBF7D0"),       # verde
    "REVISAR": PatternFill("solid", fgColor="FEF3C7"),  # amarillo
    "ERROR": PatternFill("solid", fgColor="FECACA"),    # rojo
    "FALLO": PatternFill("solid", fgColor="E5E7EB"),    # gris
}


def build_summary_report(rows: list["BatchRow"]) -> bytes:
    wb = Workbook()

    # --- Resumen ---
    ws = wb.active
    ws.title = "Resumen"
    cols = ["Archivo", "Estado", "Tipo", "Número", "Emisor", "RUC Emisor",
            "Doc Cliente", "Moneda", "Subtotal", "IGV", "Total",
            "Confianza", "Errores", "Warnings"]
    _header(ws, cols)
    for r, row in enumerate(rows, start=2):
        ws.cell(r, 1, row.filename)
        st = ws.cell(r, 2, row.status)
        st.fill = _STATUS_FILL.get(row.status, _STATUS_FILL["FALLO"])
        st.font = Font(bold=True)
        ws.cell(r, 3, row.document_type)
        ws.cell(r, 4, row.full_number)
        ws.cell(r, 5, row.issuer_name)
        ws.cell(r, 6, row.issuer_ruc)
        ws.cell(r, 7, row.customer_doc)
        ws.cell(r, 8, row.currency)
        ws.cell(r, 9, row.subtotal)
        ws.cell(r, 10, row.igv)
        ws.cell(r, 11, row.total)
        ws.cell(r, 12, row.confidence)
        ws.cell(r, 13, row.n_errors)
        ws.cell(r, 14, row.n_warnings)
    ws.freeze_panes = "A2"
    _autosize(ws)

    # --- Validación ---
    wv = wb.create_sheet("Validación")
    _header(wv, ["Archivo", "Campo", "Severidad", "Mensaje"])
    r = 2
    for row in rows:
        if row.status == "FALLO":
            wv.cell(r, 1, row.filename)
            wv.cell(r, 2, "-")
            sev = wv.cell(r, 3, "fallo")
            sev.fill = _STATUS_FILL["FALLO"]
            wv.cell(r, 4, row.error_detail)
            r += 1
            continue
        for issue in row.issues:
            wv.cell(r, 1, row.filename)
            wv.cell(r, 2, issue["field"])
            sev = wv.cell(r, 3, issue["severity"])
            sev.fill = _STATUS_FILL["ERROR"] if issue["severity"] == "error" else _STATUS_FILL["REVISAR"]
            wv.cell(r, 4, issue["message"])
            r += 1
    if r == 2:
        wv.cell(2, 1, "Sin observaciones en todo el lote ✓")
    wv.freeze_panes = "A2"
    _autosize(wv)

    # --- Totales ---
    wt = wb.create_sheet("Totales")
    n = len(rows)
    by_status = {s: sum(1 for x in rows if x.status == s) for s in ("OK", "REVISAR", "ERROR", "FALLO")}
    by_type: dict[str, int] = {}
    for x in rows:
        if x.status != "FALLO":
            by_type[x.document_type] = by_type.get(x.document_type, 0) + 1
    total_ok = sum(x.total for x in rows if x.status == "OK")

    summary = [
        ("Archivos procesados", n),
        ("OK (sin observaciones)", by_status["OK"]),
        ("Para revisar (warnings)", by_status["REVISAR"]),
        ("Con errores bloqueantes", by_status["ERROR"]),
        ("Fallos de extracción", by_status["FALLO"]),
        ("", ""),
        ("Suma de totales (solo OK)", float(total_ok)),
        ("", ""),
    ]
    for label, value in summary:
        rr = wt.max_row + 1 if wt.max_row > 1 or wt.cell(1, 1).value else 1
        wt.cell(rr, 1, label).font = Font(bold=bool(label))
        wt.cell(rr, 2, value)
    wt.cell(wt.max_row + 1, 1, "Por tipo de documento:").font = Font(bold=True)
    for t, c in sorted(by_type.items()):
        wt.cell(wt.max_row + 1, 1, f"  {t}")
        wt.cell(wt.max_row, 2, c)
    _autosize(wt)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header(ws, cols: list[str]) -> None:
    for c, name in enumerate(cols, start=1):
        cell = ws.cell(1, c, name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def _autosize(ws, max_width: int = 55) -> None:
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 3, max_width)
